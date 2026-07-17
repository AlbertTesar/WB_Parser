import pandas as pd
import openpyxl

class PriceComparer:
    def __init__(self, file1, file2, output_file):
        self.file1 = file1
        self.file2 = file2
        self.output_file = output_file

    def compare_prices(self):
        df1 = pd.read_excel(self.file1)
        df2 = pd.read_excel(self.file2)
        merged_df = pd.merge(df1, df2, on='id', suffixes=('_old', '_new'))

        merged_df['Подорожание %'] = ((merged_df['Цена со скидкой_new'] - merged_df['Цена со скидкой_old']) / merged_df['Цена со скидкой_old']) * 100

        average_discount_all = merged_df['Подорожание %'].mean()
        discounted_df = merged_df[merged_df['Подорожание %'] != 0]
        average_discount_changes = discounted_df['Подорожание %'].mean() if not discounted_df.empty else 0

        min_discount = discounted_df['Подорожание %'].min()
        min_discount_item = discounted_df[discounted_df['Подорожание %'] == min_discount].iloc[0]

        # Проверяем каждое поле на наличие и возвращаем значение или 'не указано'
        name = min_discount_item.get('Наименование', 'Название не указано') if 'Наименование' in min_discount_item else 'Название не указано'
        article = min_discount_item.get('id', 'ID не указан') if 'id' in min_discount_item else 'ID не указан'
        price = min_discount_item.get('Цена со скидкой_new', 'Цена не указана') if 'Цена со скидкой_new' in min_discount_item else 'Цена не указана'
        discount_percent = min_discount_item.get('Подорожание %', 0) if 'Подорожание %' in min_discount_item else 0
        print(df1.columns)
        print(df2.columns)

        discount_info = f"{name}, Артикул: {article}, Цена: {price}, Подешевел на: {discount_percent:.2f}%"

        discounted_df.to_excel(self.output_file, index=False)
        wb = openpyxl.load_workbook(self.output_file)
        ws = wb.active

        discount_col = merged_df.columns.get_loc('Подорожание %') + 1
        ws.cell(row=1, column=discount_col + 1, value=f'Среднее изменение цен (все товары): {average_discount_all:.2f}%')
        ws.cell(row=2, column=discount_col + 1, value=f'Среднее изменение цен (только с изменениями): {average_discount_changes:.2f}%')
        ws.cell(row=3, column=discount_col + 1, value=f'Самый подешевевший товар: {discount_info}')

        wb.save(self.output_file)
        print(f'Результаты с обновленной информацией сохранены в {self.output_file}')
        print(discount_info)
        return average_discount_all, average_discount_changes, discount_info

if __name__ == '__main__':
    comparer = PriceComparer(r'path\to\file1.xlsx', r'path\to\file2.xlsx', 'result_comparison.xlsx')
    results = comparer.compare_prices()
    print(results)
